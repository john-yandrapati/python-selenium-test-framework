"""This class is for Excel"""
from src.main.python.com.home.helper.helperInterface import iHelperInterface as HI
from openpyxl import load_workbook


class ExcelHelperClass(HI):

    def __init__(self, fileHelper, config):
        self.fileHelper = fileHelper
        self.config = config

    keyName = ""

    def dummyMethod(self):
        pass

    def getTestData(self,sheetName):

        try:
            wb = load_workbook(self.fileHelper.getExcelFilePath())
            sheet = wb[sheetName]

            listData = []
            for column in sheet.iter_cols(min_row=sheet.min_row,max_row=sheet.max_row,min_col=sheet.min_column,max_col=sheet.max_column, values_only=True):
                for i in column:
                    if i != None:
                        listData.append(i)

            return listData

        except:
            raise Exception("Unable to load excel file")